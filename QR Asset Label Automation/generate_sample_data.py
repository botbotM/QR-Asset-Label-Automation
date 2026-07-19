import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

random.seed(42)

part_names = ["MB", "SSD", "SATA", "GB", "D"]

# 重複しない末尾5桁を100個作成
set_ids = random.sample(range(10000, 100000), 100)

machine_rows = []
part_rows = []

for set_id_number in set_ids:
    set_id = str(set_id_number).zfill(5)

    # マシンを1件作成
    machine_name = random.choice(["VO1", "VO2", "VO3"])
    machine_number = f"{machine_name}{set_id}"

    machine_rows.append([machine_number])

    # 5種類のパーツを作成
    for part_name in part_names:
        random_text = "".join(
            random.choices(
                "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789",
                k=8
            )
        )

        # 必ず末尾がセットIDになる
        part_number = f"{random_text}{set_id}"

        part_rows.append([
            part_number,
            part_name
        ])
        
# -----------------------------------
# 意図的な異常データを作る
# -----------------------------------

error_records = []

# 異常を入れるセットIDを重複なしで選ぶ
error_set_ids = random.sample(set_ids, 10)

single_missing_ids = error_set_ids[:5]
double_missing_ids = error_set_ids[5:8]
duplicate_ids = error_set_ids[8:10]


def remove_part(target_set_id, target_part_name):
    """指定セットから指定パーツを1件削除する"""

    global part_rows

    set_id = str(target_set_id).zfill(5)

    before_count = len(part_rows)

    part_rows = [
        row for row in part_rows
        if not (
            row[0].endswith(set_id)
            and row[1] == target_part_name
        )
    ]

    return before_count - len(part_rows)


# パーツ1個抜けを5セット作る
for target_set_id in single_missing_ids:
    missing_part = random.choice(part_names)
    set_id = str(target_set_id).zfill(5)

    remove_part(set_id, missing_part)

    error_records.append([
        set_id,
        "パーツ1個抜け",
        missing_part
    ])


# パーツ2個抜けを3セット作る
for target_set_id in double_missing_ids:
    missing_parts = random.sample(part_names, 2)
    set_id = str(target_set_id).zfill(5)

    for missing_part in missing_parts:
        remove_part(set_id, missing_part)

    error_records.append([
        set_id,
        "パーツ2個抜け",
        "、".join(missing_parts)
    ])


# 同じパーツの重複を2セット作る
for target_set_id in duplicate_ids:
    set_id = str(target_set_id).zfill(5)
    duplicate_part = random.choice(part_names)

    # 対象パーツを探す
    target_row = next(
        row for row in part_rows
        if row[0].endswith(set_id)
        and row[1] == duplicate_part
    )

    # 管理番号を少し変えて同じ種類を追加
    random_text = "".join(
        random.choices(
            "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789",
            k=8
        )
    )

    duplicate_number = f"{random_text}{set_id}"

    part_rows.append([
        duplicate_number,
        duplicate_part
    ])

    error_records.append([
        set_id,
        "パーツ重複",
        duplicate_part
    ])

# それぞれ別々に並び替える
random.shuffle(machine_rows)
random.shuffle(part_rows)

# それぞれ別々に並び替える
random.shuffle(machine_rows)
random.shuffle(part_rows)

# -----------------------------------
# Excelファイルを作成
# -----------------------------------

workbook = Workbook()

machine_sheet = workbook.active
machine_sheet.title = "マシン"

parts_sheet = workbook.create_sheet("パーツ")
error_sheet = workbook.create_sheet("異常データ一覧")

# 見出し
machine_sheet.append([
    "管理番号"
])

parts_sheet.append([
    "管理番号",
    "パーツ名"
])

error_sheet.append([
    "末尾5桁",
    "異常の種類",
    "対象パーツ"
])

# マシンを書き込む
for row in machine_rows:
    machine_sheet.append(row)

# パーツを書き込む
for row in part_rows:
    parts_sheet.append(row)

# 異常データの正解を書き込む
for row in error_records:
    error_sheet.append(row)

# -----------------------------------
# 見た目を整える
# -----------------------------------

header_fill = PatternFill(
    fill_type="solid",
    fgColor="4472C4"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

for sheet in [
    machine_sheet,
    parts_sheet,
    error_sheet
]:
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 25

parts_sheet.column_dimensions["B"].width = 15
error_sheet.column_dimensions["B"].width = 20
error_sheet.column_dimensions["C"].width = 20

# 保存
output_name = "sample_assets.xlsx"
workbook.save(output_name)

print(f"{output_name}を作成しました")
print(f"マシン：{len(machine_rows)}件")
print(f"パーツ：{len(part_rows)}件")
print(f"異常セット：{len(error_records)}件")

management_number = "VO112345"
set_id = management_number[-5:]

print(set_id)