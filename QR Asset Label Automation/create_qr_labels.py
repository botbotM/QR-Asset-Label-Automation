from openpyxl import load_workbook
import qrcode
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

workbook = load_workbook("sample_assets.xlsx")
machine_sheet = workbook["マシン"]

machine_list = []

for row in machine_sheet.iter_rows(
    min_row=2,
    values_only=True
):
    management_number = row[0]
    set_id = management_number[-5:]
    machine_name = management_number[:3]

    machine_list.append({
        "管理番号": management_number,
        "マシン名": machine_name,
        "末尾5桁": set_id
    })

print(machine_list[0])
print(f"マシン件数：{len(machine_list)}件")

# パーツシートを選ぶ
parts_sheet = workbook["パーツ"]

part_list = []

for row in parts_sheet.iter_rows(
    min_row=2,
    values_only=True
):
    management_number = row[0]
    part_name = row[1]
    set_id = management_number[-5:]

    part_list.append({
        "管理番号": management_number,
        "パーツ名": part_name,
        "末尾5桁": set_id
    })

print(part_list[0])
print(f"パーツ件数：{len(part_list)}件")

# マシンとパーツを末尾5桁でまとめる
asset_sets = []

for machine in machine_list:
    matching_parts = []

    for part in part_list:
        if machine["末尾5桁"] == part["末尾5桁"]:
            matching_parts.append(part)

    asset_sets.append({
        "マシン": machine,
        "パーツ": matching_parts
    })

# 最初の1セットを確認
first_set = asset_sets[0]

print("マシン")
print(first_set["マシン"])

print("対応パーツ")
for part in first_set["パーツ"]:
    print(part)

print(
    f"対応パーツ数：{len(first_set['パーツ'])}件"
)

error_sets = []

for asset_set in asset_sets:
    parts = asset_set["パーツ"]

    if len(parts) != 5:
        error_sets.append(asset_set)

        machine = asset_set["マシン"]
        set_id = machine["末尾5桁"]

        part_names = []

        for part in parts:
            part_names.append(part["パーツ名"])

        print(
            f"末尾5桁：{set_id} "
            f"パーツ数：{len(parts)} "
            f"内容：{part_names}"
        )

print(f"異常セット数：{len(error_sets)}件")

# 正しいパーツ構成
expected_part_names = [
    "MB",
    "SSD",
    "SATA",
    "GB",
    "D"
]

detailed_errors = []

for asset_set in asset_sets:
    machine = asset_set["マシン"]
    parts = asset_set["パーツ"]

    set_id = machine["末尾5桁"]

    # 実際に入っているパーツ名を集める
    actual_part_names = []

    for part in parts:
        actual_part_names.append(
            part["パーツ名"]
        )

    missing_parts = []
    duplicate_parts = []
    unexpected_parts = []

    # 必要な5種類を確認
    for expected_name in expected_part_names:
        part_count = actual_part_names.count(
            expected_name
        )

        if part_count == 0:
            missing_parts.append(expected_name)

        elif part_count > 1:
            duplicate_parts.append(expected_name)

    # 想定外のパーツ名を確認
    for actual_name in actual_part_names:
        if actual_name not in expected_part_names:
            if actual_name not in unexpected_parts:
                unexpected_parts.append(actual_name)

    # 何か問題があれば異常として保存
    if missing_parts or duplicate_parts or unexpected_parts:
        detailed_errors.append({
            "末尾5桁": set_id,
            "不足": missing_parts,
            "重複": duplicate_parts,
            "想定外": unexpected_parts
        })

        print(
            f"末尾5桁：{set_id} "
            f"不足：{missing_parts} "
            f"重複：{duplicate_parts} "
            f"想定外：{unexpected_parts}"
        )

print(
    f"詳細チェックの異常セット数："
    f"{len(detailed_errors)}件"
)

# 異常セットの末尾5桁を集める
error_set_ids = []

for error in detailed_errors:
    error_set_ids.append(
        error["末尾5桁"]
    )

# 異常ではないセットを集める
normal_sets = []

for asset_set in asset_sets:
    set_id = asset_set["マシン"]["末尾5桁"]

    if set_id not in error_set_ids:
        normal_sets.append(asset_set)

print(f"正常セット数：{len(normal_sets)}件")
print(f"異常セット数：{len(detailed_errors)}件")

label_data = []

for asset_set in normal_sets:
    machine = asset_set["マシン"]
    parts = asset_set["パーツ"]

    # マシン用シール
    label_data.append({
        "種類": "マシン",
        "表示名": machine["マシン名"],
        "末尾5桁": machine["末尾5桁"],
        "QRデータ": machine["管理番号"]
    })

    # パーツ用シール
    for part in parts:
        label_data.append({
            "種類": "パーツ",
            "表示名": part["パーツ名"],
            "末尾5桁": part["末尾5桁"],
            "QRデータ": part["管理番号"]
        })

print(f"シールデータ数：{len(label_data)}件")

# 最初の6枚を確認
for label in label_data[:6]:
    print(label)
    

# 最初のシールデータを選ぶ
first_label = label_data[0]

# QRに入れる管理番号
qr_text = first_label["QRデータ"]

# QRコード画像を作る
qr_image = qrcode.make(qr_text)

# PNG画像として保存
qr_image.save("test_qr.png")

print("QRコードを作成しました")
print(f"QRの内容：{qr_text}")

# -----------------------------------
# テスト用シールを1枚作る
# -----------------------------------

first_label = label_data[0]

display_name = first_label["表示名"]
set_id = first_label["末尾5桁"]
qr_text = first_label["QRデータ"]

# QRコードを作成
qr_image = qrcode.make(qr_text).convert("RGB")

# QRはぼかさずに拡大する
qr_image = qr_image.resize(
    (220, 220),
    Image.Resampling.NEAREST
)

# シールの白い台紙
label_image = Image.new(
    "RGB",
    (600, 300),
    "white"
)

draw = ImageDraw.Draw(label_image)

# フォントを準備
font_path = "/System/Library/Fonts/Helvetica.ttc"

try:
    name_font = ImageFont.truetype(
        font_path,
        70
    )
    id_font = ImageFont.truetype(
        font_path,
        48
    )
except OSError:
    name_font = ImageFont.load_default()
    id_font = ImageFont.load_default()

# シールの枠
draw.rounded_rectangle(
    (5, 5, 595, 295),
    radius=15,
    outline="black",
    width=4
)

# 名称と末尾5桁
draw.text(
    (40, 65),
    display_name,
    fill="black",
    font=name_font
)

draw.text(
    (40, 165),
    set_id,
    fill="black",
    font=id_font
)

# QRコードを右側に配置
label_image.paste(
    qr_image,
    (340, 40)
)

label_image.save("test_label.png")

print("テスト用シールを作成しました")

# -----------------------------------
# A4印刷用PDFを作成
# -----------------------------------

output_pdf = "portfolio_qr_labels.pdf"

pdf = canvas.Canvas(
    output_pdf,
    pagesize=A4
)

page_width, page_height = A4

# 3列×8段
columns = 3
rows = 8
labels_per_page = columns * rows

# ページの余白
margin_x = 10 * mm
margin_y = 10 * mm

# シール1枚分の大きさ
label_width = (
    page_width - margin_x * 2
) / columns

label_height = (
    page_height - margin_y * 2
) / rows

qr_size = 24 * mm


def draw_label(pdf, label, x, y):
    """シールを1枚描画する"""

    # シールの外枠
    pdf.setStrokeColorRGB(0.65, 0.65, 0.65)
    pdf.setLineWidth(0.5)

    pdf.roundRect(
        x,
        y,
        label_width,
        label_height,
        2 * mm,
        stroke=1,
        fill=0
    )

    # 表示名
    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont("Helvetica-Bold", 16)

    pdf.drawString(
        x + 4 * mm,
        y + label_height - 12 * mm,
        label["表示名"]
    )

    # 末尾5桁
    pdf.setFont("Helvetica", 13)

    pdf.drawString(
        x + 4 * mm,
        y + 8 * mm,
        label["末尾5桁"]
    )

    # QRコードをメモリ上で作成
    qr_image = qrcode.make(
        label["QRデータ"]
    )

    qr_buffer = BytesIO()
    qr_image.save(
        qr_buffer,
        format="PNG"
    )
    qr_buffer.seek(0)

    qr_reader = ImageReader(qr_buffer)

    # QRを右側中央に配置
    qr_x = (
        x + label_width
        - qr_size
        - 3 * mm
    )

    qr_y = (
        y
        + (label_height - qr_size) / 2
    )

    pdf.drawImage(
        qr_reader,
        qr_x,
        qr_y,
        width=qr_size,
        height=qr_size,
        preserveAspectRatio=True,
        mask="auto"
    )


for index, label in enumerate(label_data):
    # 24枚描画したら次のページへ
    if index > 0 and index % labels_per_page == 0:
        pdf.showPage()

    position_on_page = index % labels_per_page

    column = position_on_page % columns
    row = position_on_page // columns

    x = margin_x + column * label_width

    # PDFの座標は左下が基準なので上から計算
    y = (
        page_height
        - margin_y
        - (row + 1) * label_height
    )

    draw_label(
        pdf,
        label,
        x,
        y
    )

pdf.save()

print(f"{output_pdf}を作成しました")
print(f"シール枚数：{len(label_data)}枚")
print(
    f"ページ数："
    f"{(len(label_data) + labels_per_page - 1) // labels_per_page}ページ"
)